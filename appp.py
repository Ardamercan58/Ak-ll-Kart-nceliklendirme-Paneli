import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, timedelta

import plotly.express as px
import plotly.graph_objects as go

# =============================================================================
# SAYFA YAPILANDIRMASI
# =============================================================================
st.set_page_config(
    page_title="THY Teknik ÜPK | Akıllı Önceliklendirme",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# ÖZEL STİL (THY Kurumsal Kimlik: Kırmızı / Lacivert / Beyaz)
# =============================================================================
THY_RED = "#E81932"
THY_NAVY = "#0A2240"
THY_GRAY = "#F4F6F8"

THY_TEXT = "#1A1A2E"

st.markdown(
    f"""
    <style>
        /* ---- GENEL ZEMİN VE VARSAYILAN YAZI RENGİ (tarayıcı/koyu tema fark etmeksizin) ---- */
        html, body, [class*="css"] {{
            color: {THY_TEXT} !important;
        }}
        .stApp {{
            background-color: {THY_GRAY} !important;
        }}
        .main .block-container {{
            color: {THY_TEXT} !important;
        }}

        /* ---- ANA İÇERİKTEKİ TÜM BAŞLIK / METİN ÖĞELERİ ---- */
        .main h1, .main h2, .main h3, .main h4, .main h5, .main h6,
        .main p, .main span, .main label, .main li,
        .main .stMarkdown, .main .stCaption, [data-testid="stCaptionContainer"] {{
            color: {THY_TEXT} !important;
        }}

        /* ---- METRIC (KPI) BİLEŞENLERİ ---- */
        div[data-testid="stMetricValue"] {{
            color: {THY_NAVY} !important;
        }}
        div[data-testid="stMetricLabel"] {{
            color: #4A5568 !important;
        }}
        div[data-testid="stMetricDelta"] {{
            color: {THY_TEXT} !important;
        }}

        /* ---- SEKMELER (TABS) ---- */
        .stTabs [data-baseweb="tab"] {{
            font-weight: 600;
            color: {THY_TEXT} !important;
        }}
        .stTabs [aria-selected="true"] {{
            color: {THY_RED} !important;
        }}
        .stTabs [data-baseweb="tab-panel"] {{
            color: {THY_TEXT} !important;
        }}

        /* ---- EXPANDER (Zaman Baskısı İnce Ayarı vb.) ---- */
        [data-testid="stExpander"] summary {{
            color: {THY_TEXT} !important;
        }}
        [data-testid="stExpander"] * {{
            color: {THY_TEXT} !important;
        }}

        /* ---- DATA EDITOR / DATAFRAME TABLOLARI ---- */
        [data-testid="stDataFrame"] * , [data-testid="stDataEditor"] * {{
            color: {THY_TEXT} !important;
        }}

        /* ---- BİLGİ / UYARI KUTULARI (st.info, st.warning) ---- */
        [data-testid="stAlert"] * {{
            color: {THY_TEXT} !important;
        }}

        /* ---- İNDİRME BUTONLARI VE DİĞER BUTONLAR ---- */
        .stDownloadButton button, .stButton button {{
            color: {THY_NAVY} !important;
            background-color: #FFFFFF !important;
            border: 1px solid {THY_NAVY} !important;
        }}
        .stDownloadButton button:hover, .stButton button:hover {{
            color: #FFFFFF !important;
            background-color: {THY_RED} !important;
            border: 1px solid {THY_RED} !important;
        }}

        /* ---- SIDEBAR: LACİVERT ZEMİN, BEYAZ YAZI (SABİT) ---- */
        [data-testid="stSidebar"] {{
            background-color: {THY_NAVY} !important;
        }}
        [data-testid="stSidebar"] * {{
            color: #FFFFFF !important;
        }}
        [data-testid="stSidebar"] .stSlider > div > div > div > div {{
            background-color: {THY_NAVY} !important;
        }}
        /* Sidebar içindeki input/selectbox kutularının içi okunaklı olsun */
        [data-testid="stSidebar"] input, [data-testid="stSidebar"] textarea {{
            color: {THY_TEXT} !important;
            background-color: #FFFFFF !important;
        }}
        [data-testid="stSidebar"] [data-baseweb="select"] > div {{
            background-color: #FFFFFF !important;
            color: {THY_TEXT} !important;
        }}

        /* ---- SIDEBAR BUTONLARI: "Profili Uygula" / "Sıfırla" — okunaklı koyu yazı ---- */
        [data-testid="stSidebar"] .stButton button {{
            color: {THY_TEXT} !important;
            background-color: #FFFFFF !important;
            border: 1px solid #FFFFFF !important;
            font-weight: 600;
        }}
        [data-testid="stSidebar"] .stButton button p,
        [data-testid="stSidebar"] .stButton button span,
        [data-testid="stSidebar"] .stButton button div {{
            color: {THY_TEXT} !important;
        }}
        [data-testid="stSidebar"] .stButton button:hover {{
            color: #FFFFFF !important;
            background-color: {THY_RED} !important;
            border: 1px solid {THY_RED} !important;
        }}
        [data-testid="stSidebar"] .stButton button:hover p,
        [data-testid="stSidebar"] .stButton button:hover span,
        [data-testid="stSidebar"] .stButton button:hover div {{
            color: #FFFFFF !important;
        }}

        /* ---- SLIDER DEĞER BALONU: kırmızı kutuyu kaldır, okunaklı beyaz kutu yap ---- */
        [data-testid="stThumbValue"] {{
            background-color: #FFFFFF !important;
            color: {THY_TEXT} !important;
            border: 1px solid {THY_NAVY} !important;
            box-shadow: none !important;
            font-weight: 600;
        }}
        [data-testid="stSliderTickBarMin"], [data-testid="stSliderTickBarMax"] {{
            color: #FFFFFF !important;
        }}

        /* ---- ÖZEL BLOKLAR ---- */
        div.block-container {{
            padding-top: 1.5rem;
        }}
        .thy-header {{
            background: linear-gradient(90deg, {THY_NAVY} 0%, #14345E 100%);
            padding: 22px 30px;
            border-radius: 12px;
            border-left: 8px solid {THY_RED};
            margin-bottom: 18px;
        }}
        .thy-header h1 {{
            color: #FFFFFF !important;
            margin: 0;
            font-size: 28px;
        }}
        .thy-header p {{
            color: #C9D4E3 !important;
            margin: 4px 0 0 0;
            font-size: 14px;
        }}
        .kpi-card {{
            background: #FFFFFF;
            border-radius: 12px;
            padding: 16px 18px;
            box-shadow: 0 1px 4px rgba(10,34,64,0.12);
            border-top: 4px solid {THY_RED};
        }}
        .alert-banner {{
            background-color: #FDECEA;
            border-left: 6px solid {THY_RED};
            padding: 12px 18px;
            border-radius: 8px;
            color: {THY_NAVY} !important;
            font-weight: 600;
            margin-bottom: 14px;
        }}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <div class="thy-header">
        <h1>✈️ THY Teknik ÜPK — Akıllı İş Kartı Önceliklendirme Paneli</h1>
        <p>Uçak bakım operasyonlarında kritik yol, AOG riski, parça durumu ve kaynak müsaitliğine göre
        dinamik, tam yapılandırılabilir önceliklendirme sistemi.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# =============================================================================
# SABİTLER
# =============================================================================
AIRCRAFT_TYPES = ["A320-200", "A321neo", "A330-300", "B737-800", "B777-300ER", "B787-9"]
TAIL_NUMBERS = ["TC-JFP", "TC-LNC", "TC-JRO", "TC-LGA", "TC-JJU", "TC-JOC", "TC-LOF", "TC-JJT"]
STATIONS = ["IST Hangar-1", "IST Hangar-2", "IST Line Maintenance", "SAW Hangar", "ESB Hangar"]
SHIFTS = ["Gündüz (08-16)", "Akşam (16-24)", "Gece (00-08)"]
TECHNICIANS = [
    "A. Yılmaz", "B. Demir", "C. Kaya", "D. Şahin", "E. Çelik",
    "F. Aydın", "G. Arslan", "H. Doğan", "İ. Kılıç", "J. Aksoy",
]
SKILLS = ["Avyonik", "Mekanik", "Motor", "Yapısal (Structures)", "Elektrik", "NDT"]
JOB_STATUS = ["Beklemede", "Devam Ediyor", "Parça Bekleniyor", "Kontrol Aşamasında"]
WORK_SCOPES = [
    "Alet / Avyonik Sistem Kontrolü", "Kabin İçi Koltuk / Galley Onarımı",
    "Fren Diski ve İniş Takımı Değişimi", "Hidrolik Sızıntı Testi",
    "Motor Fan Bıçağı Muayenesi (NDT)", "Korozyon Temizliği ve Boya",
    "Uçuş Kontrol Yüzeyleri Kalibrasyonu", "Oksijen Sistemi Basınç Testi",
    "APU Performans Testi", "Yakıt Sistemi Sızıntı Kontrolü",
    "İniş Takımı Yağlama ve Muayenesi", "Radom / Anten Kontrolü",
]
PART_STATUS_OPTIONS = ["Hazır", "Siparişte", "Karantina / Testte"]

WEIGHT_PRESETS = {
    "⚖️ Dengeli (Varsayılan)": dict(critical=35, aog=30, part=20, dep=3, time_gain=100, time_exp=1.0),
    "🚨 AOG Odaklı": dict(critical=25, aog=50, part=15, dep=2, time_gain=140, time_exp=1.3),
    "🛤️ Kritik Yol Odaklı": dict(critical=50, aog=20, part=15, dep=4, time_gain=80, time_exp=0.8),
    "📦 Parça Kısıtlı Operasyon": dict(critical=25, aog=20, part=40, dep=2, time_gain=90, time_exp=1.0),
    "⏱️ Zaman Baskılı (Hangar Çıkışı Yakın)": dict(critical=25, aog=25, part=15, dep=2, time_gain=180, time_exp=1.6),
}

# =============================================================================
# MOCK VERİ ÜRETİCİ
# =============================================================================
@st.cache_data
def generate_mock_data(n_jobs: int = 60, seed: int = 42) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    now = datetime(2026, 8, 13, 6, 0)

    rows = []
    for i in range(1, n_jobs + 1):
        job_id = f"JC-2026-{1000 + i}"
        aircraft = rng.choice(AIRCRAFT_TYPES)
        tail = rng.choice(TAIL_NUMBERS)
        scope = rng.choice(WORK_SCOPES)
        skill = rng.choice(SKILLS)
        station = rng.choice(STATIONS)
        shift = rng.choice(SHIFTS)
        technician = rng.choice(TECHNICIANS)
        status = rng.choice(JOB_STATUS, p=[0.35, 0.30, 0.20, 0.15])

        is_critical = int(rng.choice([1, 0], p=[0.32, 0.68]))
        is_aog_risk = int(rng.choice([1, 0], p=[0.22, 0.78]))
        part_status = rng.choice(PART_STATUS_OPTIONS, p=[0.58, 0.30, 0.12])
        dependencies = int(rng.integers(0, 7))
        planned_hours = float(rng.choice([1.5, 2.0, 3.0, 4.5, 6.0, 8.0, 12.0]))
        hours_remaining = float(rng.integers(1, 60))
        completion_pct = int(rng.integers(0, 100)) if status == "Devam Ediyor" else (
            0 if status in ("Beklemede", "Parça Bekleniyor") else int(rng.integers(70, 100))
        )
        due_at = now + timedelta(hours=hours_remaining)

        rows.append({
            "İş Kartı ID": job_id,
            "Uçak Tipi": aircraft,
            "Kuyruk Tescil": tail,
            "İş Tanımı": scope,
            "Uzmanlık Alanı": skill,
            "İstasyon": station,
            "Vardiya": shift,
            "Atanan Teknisyen": technician,
            "Durum": status,
            "Tamamlanma (%)": completion_pct,
            "Kritik Yol Mi?": is_critical,
            "AOG Riski Var mı?": is_aog_risk,
            "Parça Durumu": part_status,
            "Bağımlı İş Sayısı": dependencies,
            "Planlanan Süre (Saat)": planned_hours,
            "Hangardan Çıkışa Kalan (Saat)": hours_remaining,
            "Hedef Bitiş Zamanı": due_at,
        })

    return pd.DataFrame(rows)


# =============================================================================
# SESSION STATE — AĞIRLIKLARIN KALICI OLMASI İÇİN
# =============================================================================
defaults = WEIGHT_PRESETS["⚖️ Dengeli (Varsayılan)"]
for key, val in defaults.items():
    st.session_state.setdefault(f"w_{key}", val)

# =============================================================================
# SIDEBAR — VERİ ÖLÇEĞİ
# =============================================================================
st.sidebar.header("📦 Veri Kaynağı")
n_jobs = st.sidebar.slider("Simüle Edilecek İş Kartı Sayısı", 10, 200, 60, step=10)
raw_df = generate_mock_data(n_jobs=n_jobs)

st.sidebar.markdown("---")

# =============================================================================
# SIDEBAR — AĞIRLIK PROFİLLERİ (PRESET)
# =============================================================================
st.sidebar.header("🎯 Hazır Öncelik Profilleri")
preset_choice = st.sidebar.selectbox(
    "Bir operasyon profili seçerek ağırlıkları otomatik doldurun",
    options=list(WEIGHT_PRESETS.keys()),
    index=0,
)
col_preset1, col_preset2 = st.sidebar.columns(2)
with col_preset1:
    apply_preset = st.button("Profili Uygula", use_container_width=True)
with col_preset2:
    reset_weights = st.button("🔄 Sıfırla", use_container_width=True)

# ÖNEMLİ: Bu iki bloğun ikisi de, aşağıdaki slider'lar OLUŞTURULMADAN ÖNCE
# çalışmalı. Streamlit, bir widget key'ine (örn. "w_critical") ait slider
# aynı script çalışmasında zaten oluşturulduysa, o key için session_state
# üzerinden doğrudan değer atanmasına izin vermiyor ve hata fırlatıyor.
# Bu yüzden hem "Profili Uygula" hem "Sıfırla" mantığı sliderlardan önce yer alıyor.
if apply_preset:
    chosen = WEIGHT_PRESETS[preset_choice]
    st.session_state["w_critical"] = chosen["critical"]
    st.session_state["w_aog"] = chosen["aog"]
    st.session_state["w_part"] = chosen["part"]
    st.session_state["w_dep"] = chosen["dep"]
    st.session_state["w_time_gain"] = chosen["time_gain"]
    st.session_state["w_time_exp"] = chosen["time_exp"]
    st.rerun()

if reset_weights:
    default_preset = WEIGHT_PRESETS["⚖️ Dengeli (Varsayılan)"]
    st.session_state["w_critical"] = default_preset["critical"]
    st.session_state["w_aog"] = default_preset["aog"]
    st.session_state["w_part"] = default_preset["part"]
    st.session_state["w_dep"] = default_preset["dep"]
    st.session_state["w_time_gain"] = default_preset["time_gain"]
    st.session_state["w_time_exp"] = default_preset["time_exp"]
    st.rerun()

st.sidebar.markdown("---")

# =============================================================================
# SIDEBAR — MANUEL AĞIRLIK AYARLARI
# =============================================================================
st.sidebar.header("⚙️ Öncelik Algoritması Ağırlıkları")

w_critical = st.sidebar.slider("Kritik Yol Ağırlığı", 0, 60, key="w_critical")
w_aog = st.sidebar.slider("AOG Riski Ağırlığı", 0, 60, key="w_aog")
w_part = st.sidebar.slider("Parça Hazır Olma Ağırlığı", 0, 50, key="w_part")
w_dep = st.sidebar.slider("Bağımlı İş Çarpanı", 0, 8, key="w_dep")

with st.sidebar.expander("⏱️ Zaman Baskısı İnce Ayarı"):
    w_time_gain = st.slider("Zaman Baskısı Katsayısı (sabit pay)", 20, 300, key="w_time_gain")
    w_time_exp = st.slider("Zaman Baskısı Üssü (yaklaşma etkisini keskinleştirir)", 0.5, 2.5, key="w_time_exp", step=0.1)

st.sidebar.markdown("---")

# =============================================================================
# SIDEBAR — FİLTRELER
# =============================================================================
st.sidebar.header("🔍 Filtreleme ve Arama")

search_text = st.sidebar.text_input("İş Kartı ID / Kuyruk / İş Tanımı Ara", "")

selected_aircraft = st.sidebar.multiselect(
    "Uçak Tipi", options=AIRCRAFT_TYPES, default=AIRCRAFT_TYPES
)
selected_stations = st.sidebar.multiselect(
    "İstasyon / Hangar", options=STATIONS, default=STATIONS
)
selected_shifts = st.sidebar.multiselect(
    "Vardiya", options=SHIFTS, default=SHIFTS
)
selected_status = st.sidebar.multiselect(
    "İş Durumu", options=JOB_STATUS, default=JOB_STATUS
)

part_filter = st.sidebar.radio(
    "Parça Durumu Filtresi", ["Tümü", "Sadece Parçası Hazır Olanlar", "Sadece Bekleyenler"]
)

col_f1, col_f2 = st.sidebar.columns(2)
with col_f1:
    only_critical = st.checkbox("Sadece Kritik Yol", value=False)
with col_f2:
    only_aog = st.checkbox("Sadece AOG Riski", value=False)

hours_range = st.sidebar.slider(
    "Hangardan Çıkışa Kalan Süre (Saat) Aralığı",
    min_value=0, max_value=int(raw_df["Hangardan Çıkışa Kalan (Saat)"].max()) + 1,
    value=(0, int(raw_df["Hangardan Çıkışa Kalan (Saat)"].max()) + 1),
)

top_n = st.sidebar.slider("Genel Bakışta Gösterilecek En Öncelikli İş Sayısı", 5, 30, 10)

# =============================================================================
# ÖNCELİK SKORU HESAPLAMA MOTORU
# =============================================================================
def calculate_priority_scores(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        df["Öncelik Skoru"] = []
        return df

    part_score_map = {"Hazır": 1.0, "Karantina / Testte": 0.3, "Siparişte": 0.0}

    time_component = w_time_gain / np.power(df["Hangardan Çıkışa Kalan (Saat)"] + 1, w_time_exp)

    scores = (
        (df["Kritik Yol Mi?"] * w_critical)
        + (df["AOG Riski Var mı?"] * w_aog)
        + (df["Parça Durumu"].map(part_score_map) * w_part)
        + (df["Bağımlı İş Sayısı"] * w_dep)
        + time_component
    )
    df = df.copy()
    df["Öncelik Skoru"] = np.round(scores, 1)

    max_score = df["Öncelik Skoru"].max()
    if max_score > 0:
        df["Öncelik (Normalize 0-100)"] = np.round(df["Öncelik Skoru"] / max_score * 100, 1)
    else:
        df["Öncelik (Normalize 0-100)"] = 0.0

    def tier(score_norm):
        if score_norm >= 75:
            return "🔴 Acil"
        elif score_norm >= 45:
            return "🟠 Yüksek"
        elif score_norm >= 20:
            return "🟡 Orta"
        return "🟢 Düşük"

    df["Öncelik Seviyesi"] = df["Öncelik (Normalize 0-100)"].apply(tier)

    return df.sort_values(by="Öncelik Skoru", ascending=False).reset_index(drop=True)


# =============================================================================
# FİLTRELERİ UYGULA
# =============================================================================
filtered_df = raw_df[
    raw_df["Uçak Tipi"].isin(selected_aircraft)
    & raw_df["İstasyon"].isin(selected_stations)
    & raw_df["Vardiya"].isin(selected_shifts)
    & raw_df["Durum"].isin(selected_status)
    & raw_df["Hangardan Çıkışa Kalan (Saat)"].between(hours_range[0], hours_range[1])
].copy()

if part_filter == "Sadece Parçası Hazır Olanlar":
    filtered_df = filtered_df[filtered_df["Parça Durumu"] == "Hazır"]
elif part_filter == "Sadece Bekleyenler":
    filtered_df = filtered_df[filtered_df["Parça Durumu"] != "Hazır"]

if only_critical:
    filtered_df = filtered_df[filtered_df["Kritik Yol Mi?"] == 1]
if only_aog:
    filtered_df = filtered_df[filtered_df["AOG Riski Var mı?"] == 1]

if search_text.strip():
    q = search_text.strip().lower()
    mask = (
        filtered_df["İş Kartı ID"].str.lower().str.contains(q)
        | filtered_df["Kuyruk Tescil"].str.lower().str.contains(q)
        | filtered_df["İş Tanımı"].str.lower().str.contains(q)
    )
    filtered_df = filtered_df[mask]

processed_df = calculate_priority_scores(filtered_df)

# =============================================================================
# KRİTİK UYARI BANDI
# =============================================================================
urgent_mask = (
    (processed_df["AOG Riski Var mı?"] == 1)
    & (processed_df["Hangardan Çıkışa Kalan (Saat)"] <= 6)
) if not processed_df.empty else pd.Series(dtype=bool)

n_urgent = int(urgent_mask.sum()) if not processed_df.empty else 0
if n_urgent > 0:
    st.markdown(
        f"""<div class="alert-banner">⚠️ DİKKAT: {n_urgent} iş kartı AOG riski taşıyor ve
        6 saatten az süre içinde hangar çıkışı bekleniyor. Bu işler acil müdahale gerektiriyor.</div>""",
        unsafe_allow_html=True,
    )

# =============================================================================
# SEKMELER
# =============================================================================
tab_overview, tab_table, tab_analytics, tab_settings = st.tabs(
    ["📊 Genel Bakış", "📋 İş Kartları", "📈 Analitik", "🛠️ Ayarlar & Metodoloji"]
)

# -----------------------------------------------------------------------------
# TAB 1 — GENEL BAKIŞ
# -----------------------------------------------------------------------------
with tab_overview:
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        st.markdown('<div class="kpi-card">', unsafe_allow_html=True)
        st.metric("Toplam İş Kartı", len(processed_df))
        st.markdown('</div>', unsafe_allow_html=True)
    with c2:
        st.markdown('<div class="kpi-card">', unsafe_allow_html=True)
        st.metric("Kritik Yoldaki İşler", int(processed_df["Kritik Yol Mi?"].sum()) if not processed_df.empty else 0)
        st.markdown('</div>', unsafe_allow_html=True)
    with c3:
        st.markdown('<div class="kpi-card">', unsafe_allow_html=True)
        st.metric("AOG Riski Taşıyanlar", int(processed_df["AOG Riski Var mı?"].sum()) if not processed_df.empty else 0)
        st.markdown('</div>', unsafe_allow_html=True)
    with c4:
        parca_bekleyen = int((processed_df["Parça Durumu"] != "Hazır").sum()) if not processed_df.empty else 0
        st.markdown('<div class="kpi-card">', unsafe_allow_html=True)
        st.metric("Parça Bekleyen İşler", parca_bekleyen)
        st.markdown('</div>', unsafe_allow_html=True)
    with c5:
        toplam_saat = round(processed_df["Planlanan Süre (Saat)"].sum(), 1) if not processed_df.empty else 0
        st.markdown('<div class="kpi-card">', unsafe_allow_html=True)
        st.metric("Toplam Planlanan İşçilik (Saat)", toplam_saat)
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("####")
    left, right = st.columns([1.4, 1])

    with left:
        st.subheader(f"🏁 En Öncelikli {top_n} İş Kartı")
        if processed_df.empty:
            st.info("Seçili filtrelere uyan iş kartı bulunamadı.")
        else:
            top_df = processed_df.head(top_n)
            fig = px.bar(
                top_df[::-1],
                x="Öncelik Skoru",
                y="İş Kartı ID",
                orientation="h",
                color="Öncelik Seviyesi",
                color_discrete_map={
                    "🔴 Acil": THY_RED, "🟠 Yüksek": "#FF8C42",
                    "🟡 Orta": "#FFC857", "🟢 Düşük": "#4CAF50",
                },
                hover_data=["Uçak Tipi", "Kuyruk Tescil", "İş Tanımı", "Atanan Teknisyen"],
            )
            fig.update_layout(
                height=max(320, 26 * len(top_df)),
                margin=dict(l=0, r=10, t=10, b=0),
                plot_bgcolor="white",
                legend_title_text="Öncelik Seviyesi",
            )
            st.plotly_chart(fig, use_container_width=True)

    with right:
        st.subheader("📶 Öncelik Seviyesi Dağılımı")
        if processed_df.empty:
            st.info("Veri yok.")
        else:
            tier_counts = processed_df["Öncelik Seviyesi"].value_counts().reindex(
                ["🔴 Acil", "🟠 Yüksek", "🟡 Orta", "🟢 Düşük"]
            ).fillna(0)
            fig2 = go.Figure(
                data=[go.Pie(
                    labels=tier_counts.index,
                    values=tier_counts.values,
                    hole=0.55,
                    marker=dict(colors=[THY_RED, "#FF8C42", "#FFC857", "#4CAF50"]),
                )]
            )
            fig2.update_layout(margin=dict(l=0, r=0, t=10, b=0), height=320, showlegend=True)
            st.plotly_chart(fig2, use_container_width=True)

    st.markdown("####")
    st.subheader("🛩️ Uçak Bazında Açık İş Yükü")
    if processed_df.empty:
        st.info("Veri yok.")
    else:
        by_tail = processed_df.groupby("Kuyruk Tescil").agg(
            İş_Sayısı=("İş Kartı ID", "count"),
            Ortalama_Öncelik=("Öncelik Skoru", "mean"),
            Toplam_Saat=("Planlanan Süre (Saat)", "sum"),
        ).round(1).reset_index().sort_values("Ortalama_Öncelik", ascending=False)
        fig3 = px.bar(
            by_tail, x="Kuyruk Tescil", y="Toplam_Saat", color="Ortalama_Öncelik",
            color_continuous_scale=["#4CAF50", "#FFC857", THY_RED],
            text="İş_Sayısı",
        )
        fig3.update_traces(texttemplate="%{text} iş", textposition="outside")
        fig3.update_layout(margin=dict(l=0, r=0, t=10, b=0), height=340, plot_bgcolor="white")
        st.plotly_chart(fig3, use_container_width=True)

# -----------------------------------------------------------------------------
# TAB 2 — İŞ KARTLARI (DETAYLI TABLO)
# -----------------------------------------------------------------------------
with tab_table:
    st.subheader("📋 Öncelik Sıralamasına Göre Tüm İş Kartları")
    st.caption("Hücreleri doğrudan tablo üzerinde düzenleyebilir, sütun başlıklarına tıklayarak sıralayabilirsiniz.")

    if processed_df.empty:
        st.warning("Seçili filtrelere uyan iş kartı bulunamadı. Lütfen filtreleri gözden geçirin.")
    else:
        display_cols = [
            "İş Kartı ID", "Uçak Tipi", "Kuyruk Tescil", "İş Tanımı", "Uzmanlık Alanı",
            "İstasyon", "Vardiya", "Atanan Teknisyen", "Durum", "Tamamlanma (%)",
            "Kritik Yol Mi?", "AOG Riski Var mı?", "Parça Durumu", "Bağımlı İş Sayısı",
            "Planlanan Süre (Saat)", "Hangardan Çıkışa Kalan (Saat)",
            "Öncelik Skoru", "Öncelik (Normalize 0-100)", "Öncelik Seviyesi",
        ]

        edited_df = st.data_editor(
            processed_df[display_cols],
            use_container_width=True,
            height=520,
            hide_index=True,
            disabled=[c for c in display_cols if c not in ("Durum", "Atanan Teknisyen", "Tamamlanma (%)")],
            column_config={
                "Öncelik Skoru": st.column_config.ProgressColumn(
                    "Öncelik Skoru", min_value=0,
                    max_value=float(processed_df["Öncelik Skoru"].max()) or 1.0, format="%.1f",
                ),
                "Tamamlanma (%)": st.column_config.ProgressColumn(
                    "Tamamlanma (%)", min_value=0, max_value=100, format="%d%%",
                ),
                "Durum": st.column_config.SelectboxColumn("Durum", options=JOB_STATUS),
                "Atanan Teknisyen": st.column_config.SelectboxColumn("Atanan Teknisyen", options=TECHNICIANS),
                "Kritik Yol Mi?": st.column_config.CheckboxColumn("Kritik Yol Mi?"),
                "AOG Riski Var mı?": st.column_config.CheckboxColumn("AOG Riski Var mı?"),
            },
            key="job_card_editor",
        )
        st.caption("ℹ️ Not: Bu görünümdeki manuel düzenlemeler oturum içinde geçerlidir; kaynak simülasyon verisini değiştirmez.")

# -----------------------------------------------------------------------------
# TAB 3 — ANALİTİK
# -----------------------------------------------------------------------------
with tab_analytics:
    if processed_df.empty:
        st.info("Analiz için veri bulunmuyor.")
    else:
        colA, colB = st.columns(2)

        with colA:
            st.subheader("⏳ Öncelik Skoru vs. Kalan Süre")
            fig4 = px.scatter(
                processed_df, x="Hangardan Çıkışa Kalan (Saat)", y="Öncelik Skoru",
                color="Öncelik Seviyesi", size="Planlanan Süre (Saat)",
                hover_data=["İş Kartı ID", "Kuyruk Tescil", "İş Tanımı"],
                color_discrete_map={
                    "🔴 Acil": THY_RED, "🟠 Yüksek": "#FF8C42",
                    "🟡 Orta": "#FFC857", "🟢 Düşük": "#4CAF50",
                },
            )
            fig4.update_layout(plot_bgcolor="white", height=380, margin=dict(l=0, r=0, t=10, b=0))
            st.plotly_chart(fig4, use_container_width=True)

        with colB:
            st.subheader("📦 Parça Durumu Dağılımı")
            part_counts = processed_df["Parça Durumu"].value_counts().reset_index()
            part_counts.columns = ["Parça Durumu", "Adet"]
            fig5 = px.bar(
                part_counts, x="Parça Durumu", y="Adet", color="Parça Durumu",
                color_discrete_map={"Hazır": "#4CAF50", "Siparişte": "#FFC857", "Karantina / Testte": THY_RED},
                text="Adet",
            )
            fig5.update_layout(plot_bgcolor="white", height=380, margin=dict(l=0, r=0, t=10, b=0), showlegend=False)
            st.plotly_chart(fig5, use_container_width=True)

        colC, colD = st.columns(2)

        with colC:
            st.subheader("👷 Teknisyen Bazında İş Yükü")
            by_tech = processed_df.groupby("Atanan Teknisyen").agg(
                İş_Sayısı=("İş Kartı ID", "count"),
                Toplam_Saat=("Planlanan Süre (Saat)", "sum"),
            ).reset_index().sort_values("Toplam_Saat", ascending=True)
            fig6 = px.bar(
                by_tech, x="Toplam_Saat", y="Atanan Teknisyen", orientation="h",
                text="İş_Sayısı", color="Toplam_Saat", color_continuous_scale=["#C9D4E3", THY_NAVY],
            )
            fig6.update_traces(texttemplate="%{text} iş", textposition="outside")
            fig6.update_layout(plot_bgcolor="white", height=380, margin=dict(l=0, r=0, t=10, b=0))
            st.plotly_chart(fig6, use_container_width=True)

        with colD:
            st.subheader("🏭 İstasyon Bazında Kritik/AOG Yoğunluğu")
            by_station = processed_df.groupby("İstasyon").agg(
                Kritik=("Kritik Yol Mi?", "sum"),
                AOG=("AOG Riski Var mı?", "sum"),
            ).reset_index()
            fig7 = go.Figure()
            fig7.add_bar(name="Kritik Yol", x=by_station["İstasyon"], y=by_station["Kritik"], marker_color=THY_NAVY)
            fig7.add_bar(name="AOG Riski", x=by_station["İstasyon"], y=by_station["AOG"], marker_color=THY_RED)
            fig7.update_layout(
                barmode="group", plot_bgcolor="white", height=380,
                margin=dict(l=0, r=0, t=10, b=0), legend=dict(orientation="h", y=1.1),
            )
            st.plotly_chart(fig7, use_container_width=True)

# -----------------------------------------------------------------------------
# TAB 4 — AYARLAR & METODOLOJİ + EXCEL EXPORT
# -----------------------------------------------------------------------------
with tab_settings:
    st.subheader("🧮 Öncelik Skoru Nasıl Hesaplanır?")
    st.markdown(
        f"""
Her iş kartı için öncelik skoru aşağıdaki bileşenlerin toplamıdır:

| Bileşen | Açıklama | Güncel Ağırlık |
|---|---|---|
| **Kritik Yol** | İş kritik yoldaysa tam ağırlık eklenir | **{w_critical}** |
| **AOG Riski** | Uçağı yerde bırakma riski varsa tam ağırlık eklenir | **{w_aog}** |
| **Parça Hazırlığı** | Hazır: ×1.0, Karantinada: ×0.3, Siparişte: ×0.0 | **{w_part}** |
| **Bağımlı İş Sayısı** | Bu işe bağımlı diğer iş sayısı × çarpan | **{w_dep}** |
| **Zaman Baskısı** | `{w_time_gain} / (kalan_saat + 1) ^ {w_time_exp}` — hangar çıkışı yaklaştıkça skor keskin şekilde artar | dinamik |

Nihai skor, karşılaştırma kolaylığı için ayrıca **0-100 aralığında normalize edilir** ve
dört seviyeye ayrılır: 🔴 Acil (≥75), 🟠 Yüksek (≥45), 🟡 Orta (≥20), 🟢 Düşük (<20).

Sol menüdeki **hazır profiller** (AOG Odaklı, Kritik Yol Odaklı vb.) bu ağırlıkları operasyonel
önceliğe göre tek tıkla değiştirir; ardından ağırlıkları dilediğiniz gibi ince ayar yapabilirsiniz.
        """
    )

    st.markdown("---")
    st.subheader("📥 Rapor Dışa Aktarma")

    def to_excel(df: pd.DataFrame) -> bytes:
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            export_df = df.drop(columns=["Hedef Bitiş Zamanı"], errors="ignore")
            export_df.to_excel(writer, index=False, sheet_name="Oncelikli_Is_Kartlari")

            summary = pd.DataFrame({
                "Metrik": [
                    "Toplam İş Kartı", "Kritik Yoldaki İşler", "AOG Riski Taşıyanlar",
                    "Parça Bekleyen İşler", "Toplam Planlanan Saat", "Rapor Oluşturma Zamanı",
                ],
                "Değer": [
                    len(df),
                    int(df["Kritik Yol Mi?"].sum()) if not df.empty else 0,
                    int(df["AOG Riski Var mı?"].sum()) if not df.empty else 0,
                    int((df["Parça Durumu"] != "Hazır").sum()) if not df.empty else 0,
                    round(df["Planlanan Süre (Saat)"].sum(), 1) if not df.empty else 0,
                    datetime.now().strftime("%Y-%m-%d %H:%M"),
                ],
            })
            summary.to_excel(writer, index=False, sheet_name="Ozet")

            workbook = writer.book
            ws = writer.sheets["Oncelikli_Is_Kartlari"]

            from openpyxl.styles import PatternFill, Font, Alignment
            header_fill = PatternFill(start_color="0A2240", end_color="0A2240", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if "Öncelik Skoru" in export_df.columns:
                score_col_idx = export_df.columns.get_loc("Öncelik Skoru") + 1
                red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                orange_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=score_col_idx)
                    if isinstance(cell.value, (int, float)):
                        norm_row = export_df.iloc[row_idx - 2]
                        norm_val = norm_row.get("Öncelik (Normalize 0-100)", 0)
                        if norm_val >= 75:
                            cell.fill = red_fill
                        elif norm_val >= 45:
                            cell.fill = orange_fill

            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 45)

        return output.getvalue()

    excel_data = to_excel(processed_df)
    st.download_button(
        label="📥 Önceliklendirilmiş Vardiya Raporunu İndir (Excel)",
        data=excel_data,
        file_name=f"THY_Teknik_UPK_Vardiya_Raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    csv_data = processed_df.drop(columns=["Hedef Bitiş Zamanı"], errors="ignore").to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        label="📄 Filtrelenmiş Veriyi CSV Olarak İndir",
        data=csv_data,
        file_name=f"THY_Teknik_UPK_Filtreli_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mime="text/csv",
        use_container_width=True,
    )

    st.markdown("---")
    st.caption(
        "Bu panel, gösterim amaçlı simüle edilmiş veriler kullanır. Gerçek operasyonel kullanım için "
        "İş Kartı ID, uçak durumu ve parça bilgilerinin canlı MRO / ERP sistemleriyle entegre edilmesi önerilir."
    )
